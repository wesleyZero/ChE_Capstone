import pandas as pd

def npv_graphs():
    filename = 'npv_code.xlsx'
    sheet_name = 'cf'  # Specify the sheet name or number
    data = pd.read_excel(filename, sheet_name=sheet_name)

    n = 5  # Number of times to duplicate the sheet
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        for i in range(1, n + 1):
            # Assuming 'data' is the DataFrame you want to duplicate
            data.to_excel(writer, sheet_name=f'DuplicatedSheet{i}', index=False)
    
    # For modifying specific cells, openpyxl is more suited
    from openpyxl import load_workbook
    
    workbook = load_workbook(filename=filename)
    cell_row = 10  # Row of the cell to modify
    cell_col = 'C'  # Column of the cell to modify
    new_value = 69  # New value to assign
    
    for i in range(1, n + 1):
        sheet = workbook[f'ModifiedSheet{i}'] if f'ModifiedSheet{i}' in workbook.sheetnames else workbook.create_sheet(f'ModifiedSheet{i}')
        # Make sure the data is copied to the new 'ModifiedSheet'
        for r_index, row in enumerate(data.values, start=1):
            for c_index, value in enumerate(row, start=1):
                sheet.cell(row=r_index, column=c_index, value=value)
        
        # Modify the specific cell
        sheet[f'{cell_col}{cell_row}'] = new_value
    
    workbook.save(filename)

# Remember to call the function to execute
# npv_graphs()
